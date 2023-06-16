import openpyxl
from datetime import time


def create_excel_sheet(file_path):
    workbook = openpyxl.Workbook()
    workbook.save(file_path)
    print(f"Excel sheet created successfully: {file_path}")


def write_data_to_excel(file_path, data, sheet_name):
    workbook = openpyxl.load_workbook(file_path)

    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])

    sheet = workbook.create_sheet(title=sheet_name)

    for row_index, row_data in enumerate(data, start=1):
        for col_index, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=col_index, value=cell_value)

    workbook.save(file_path)


def read_data_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))

    for row in data:
        print(row)


def get_cell_value(file_path, sheet_name, column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    max_row = sheet.max_row

    for row in range(2, max_row + 1):
        cell_value = sheet[column + str(row)].value
        print(cell_value)

    workbook.close()


def validate_time_sheet(file_path, new_sheet):
    workbook = openpyxl.load_workbook(file_path)
    if new_sheet in workbook.sheetnames:
        workbook.remove(workbook[new_sheet])
        print(f'Workbook sheet as updated{new_sheet}')

    worksheet = workbook.active

    results_sheet = workbook.create_sheet(new_sheet)

    results = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        emp_id = row[0]
        times = row[1:]

        continuous_times_below_7 = False
        weeks = []

        for i, time_value in enumerate(times, start=1):
            if isinstance(time_value, time) and time_value.hour <= 7:
                if not continuous_times_below_7:
                    week = worksheet.cell(row=1, column=i + 1).value
                    weeks.append(week)
            else:
                if continuous_times_below_7 or time_value == "-":
                    continuous_times_below_7 = False
                if time_value == "-":
                    week = worksheet.cell(row=1, column=i + 1).value
                    weeks.append(week)

        if len(weeks) >= 2 or continuous_times_below_7:
            results.append(emp_id)

    results_sheet.append(["Employee ID"])
    for emp_id in results:
        print(emp_id)
        results_sheet.append([emp_id])

    workbook.save(file_path)


def validate_and_get_email(file_path, expected_sheet, actual_sheet):
    workbook = openpyxl.load_workbook(file_path)

    results_sheet = workbook[expected_sheet]
    wk_sheet = workbook[actual_sheet]

    emp_data = {}

    for row in wk_sheet.iter_rows(min_row=2, values_only=True):
        emp_id = row[0]
        emp_emailID = row[1]
        emp_data[emp_id] = emp_emailID

    results = []

    for row in results_sheet.iter_rows(min_row=2, values_only=True):
        emp_id = row[0]
        emp_emailID = emp_data.get(emp_id)
        if emp_emailID:
            results.append(emp_emailID)

    workbook.close()

    return results
